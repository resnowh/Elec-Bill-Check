import json
import requests
from dingtalkchatbot.chatbot import DingtalkChatbot
import openpyxl
from datetime import datetime,  timedelta


def write_to_excel(remaining_amount):
    """将剩余电量和查询时间写入 Excel 表格"""
    # 打开或创建 Excel 文件
    wb = openpyxl.load_workbook('electricity_records.xlsx')
    # 获取默认的工作表
    sheet = wb.active

    # 如果不存在名为“电费记录”的工作表，则创建一个
    if '电费记录' not in wb.sheetnames:
        wb.create_sheet(title='电费记录')
        sheet = wb['电费记录']
        # 写入表头
        sheet.cell(row=1, column=1, value='剩余电量')
        sheet.cell(row=1, column=2, value='查询时间')

    # 获取当前时间
    query_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 找到第一个空行，写入数据
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=remaining_amount)
    sheet.cell(row=next_row, column=2, value=query_time)

    # 保存 Excel 文件
    wb.save('electricity_records.xlsx')

def get_yesterday_electricity_usage(remaining_amount):
    """获取昨日使用电量并计算昨日电费"""
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['电费记录']

        yesterday_records = []

        # 遍历每一条数据
        for row in range(sheet.max_row, 1, -1):
            # 获取时间和电量
            record_time = sheet.cell(row=row, column=2).value
            record_amount = sheet.cell(row=row, column=1).value

            # 如果 record_time 是字符串类型，转换为 datetime 类型
            if isinstance(record_time, str):
                record_time = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")

            # 如果记录的日期是昨天，则将其添加到列表中
            if record_time.date() == datetime.now().date() - timedelta(days=1):
                yesterday_records.append(record_amount)

            # 如果找到了昨天的第一条记录和最后一条记录，则退出循环
            if len(yesterday_records) == 2:
                break

        # 如果找到了昨天的第一条记录和最后一条记录，则计算昨日使用电量
        if len(yesterday_records) == 2:
            yesterday_usage = yesterday_records[1] - yesterday_records[0]
            if (yesterday_usage != 0 and yesterday_usage != remaining_amount):
                return yesterday_usage
            else:
                # 如果未找到符合条件的记录，则返回未找到昨日电费数据
                return "未找到昨日电费数据"
        else:
            # 如果未找到符合条件的记录，则返回未找到昨日电费数据
            return "未找到昨日电费数据"

    except FileNotFoundError:
        return "未找到电费记录文件"

def get_past24hours_electricity_usage(remaining_amount):
    """获取过去24小时内使用的电量并计算消耗电费"""
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['电费记录']

        # 遍历每一条数据
        for row in range(sheet.max_row, 1, -1):
            # 获取时间和电量
            record_time = sheet.cell(row=row, column=2).value
            record_amount = sheet.cell(row=row, column=1).value

            # 检查 record_time 是否为字符串类型，如果是则转换为 datetime 类型
            if isinstance(record_time, str):
                record_time = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")

            # 计算与当前时间的时间间隔
            time_difference = datetime.now() - record_time

            # 如果时间间隔在 24 小时到 48 小时之间，则作为昨日的记录
            if timedelta(days=1) <= time_difference < timedelta(days=2):
                # 计算昨日使用电量
                yesterday_usage = record_amount - remaining_amount
                return yesterday_usage

            # 如果时间间隔大于48小时，停止遍历
            elif time_difference >= timedelta(days=2):
                break

        # 如果遍历完所有数据仍未找到符合条件的记录，则返回未找到昨日电费数据
        return "未找到昨日电费数据"

    except FileNotFoundError:
        return "未找到电费记录文件"

def check_ifSomebodyPay(remaining_amount):
    """检查是否有人充钱"""
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['电费记录']

        #读取上一条电费记录
        record_amount = sheet.cell(row=sheet.max_row - 1, column=1).value
        record_time = sheet.cell(row=sheet.max_row - 1, column=2).value
        print(record_amount,record_time)
        # 如果钱变多了
        if (float(remaining_amount) > float(record_amount)):
            # 计算昨日使用电量
            increased_amount = float(remaining_amount) - float(record_amount)
            return increased_amount
        else:
            return 0

    except FileNotFoundError:
        print ("未找到电费记录文件")

def check_ifUsageChange(remaining_amount):
    """检查是否数据更新"""
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['电费记录']

        #读取上一条电费记录
        record_amount = sheet.cell(row=sheet.max_row - 1, column=1).value
        #record_time = sheet.cell(row=sheet.max_row - 1, column=2).value

        # 如果钱没变
        if (float(remaining_amount) == float(record_amount)):
            return False
        else:
            return True

    except FileNotFoundError:
        print ("未找到电费记录文件")

def get_electricity_bill():
    """获取电费信息"""
    url = "http://172.31.248.26:8988/web/Common/Tsm.html"
    headers = {}
    data_dict = {
        "query_elec_roominfo": {
            "aid": "0030000000007301",
            "account": "158086",
            "room": {
                "roomid": room_id,
                "room": "roomid"
            },
            "floor": {
                "floorid": "",
                "floor": ""
            },
            "area": {
                "area": "",
                "areaname": ""
            },
            "building": {
                "buildingid": "",
                "building": ""
            },
            "extdata": "info1="
        }
    }
    # 将字典转换为 JSON 字符串
    jsondata = json.dumps(data_dict)

    response = requests.post(url, headers=headers, data={"jsondata": jsondata, "funname": "synjones.onecard.query.elec.roominfo"})

    if response.status_code == 200:
        print("电费信息获取成功")
        return response.text
    else:
        print("电费信息获取失败")
        return None

def parse_electricity_bill(bill):
    """解析电费信息"""
    data = json.loads(bill)
    remaining_amount = data['query_elec_roominfo']['errmsg'].split('剩余金额:')[1]
    return float(remaining_amount)

def send_notification(remaining_amount, yesterday_usage, increased_amount):
    """发送电费通知"""
    #仅返回两位小数
    remaining_amount = round(remaining_amount, 2)
    yesterday_usage = round(yesterday_usage, 2)
    increased_amount = round(increased_amount, 2)

    xiaoding = DingtalkChatbot(webhook, secret=secret)
    text = ""

    if remaining_amount < limit:
        #钱到达阈值
        text += f"⚠️ {room} 宿舍用电即将欠费，请尽快充值"
        xiaoding.send_text(text, is_at_all=True)
    else:
        text += f"🔋【电费】{room} \n"
        # 钱变多了
        if increased_amount > 0:
            text += f"💰️有人充电费啦！电费余额增加了 {increased_amount} 元！\n"
        #正常的报告信息
        text += f"目前剩余电费 {remaining_amount} 元,\n"
        text += f"昨天使用电费 {yesterday_usage} 元。"
        xiaoding.send_text(text)


def main():
    """主函数"""
    bill = get_electricity_bill()
    if bill:
        remaining_amount = parse_electricity_bill(bill)
        print("剩余电费:", remaining_amount)
        #写入本地表格
        write_to_excel(remaining_amount)
        #读取是否存在昨日电费
        yesterday_usage = get_yesterday_electricity_usage(remaining_amount)
        print("昨日消耗电费:", yesterday_usage)
        #读取是否有人充钱
        increased_amount = check_ifSomebodyPay(remaining_amount)
        #如果数据更新，再发送通知
        if check_ifUsageChange(remaining_amount):
            send_notification(remaining_amount, yesterday_usage, increased_amount)
    print("电费检查程序结束")

if __name__ == "__main__":
    limit = 20  # 欠费预警阈值
    room = '3S527'  # 房间号
    room_id = '300352711'  # 电费查询号码
    # 机器人参数，webhook 和 secret，使用时在PC端创建机器人后可以查看并替换成自己的
    webhook = 'https://oapi.dingtalk.com/robot/send?access_token=2e1a8e3bf5c77c5d3e341b63494239d537c8a18dc76653f1231bd52dec4bcfb9'
    secret = 'SEC8834a56af271fb0246db77347726811e3aa13080b888db6c44204db7ab8c0f93'
    main()
